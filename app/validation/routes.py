import io
import os

from flask import (Blueprint, abort, jsonify, redirect, render_template,
                   request, send_file, url_for)
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app.extensions import db
from app.models import Activity, ModuleData, Project
from app.report import build_report, render_pdf
from app.validation.modules import MODULE_BY_ID, compute, sections

validation_bp = Blueprint("validation", __name__, url_prefix="/hoja")

ALLOWED_EXCEL = {".xlsx", ".xlsm"}


@validation_bp.route("/<int:project_id>")
@login_required
def menu(project_id):
    project = db.get_or_404(Project, project_id)
    return render_template("validation/menu.html",
                           project=project, sections=sections())


@validation_bp.route("/<int:project_id>/info", methods=["POST"])
@login_required
def update_info(project_id):
    project = db.get_or_404(Project, project_id)
    project.molecule = (request.form.get("molecule") or "").strip()
    project.code = (request.form.get("code") or "").strip()
    project.unit = request.form.get("unit") or "ng/mL"
    try:
        project.decimals = max(0, min(10, int(request.form.get("decimals", 4))))
    except (TypeError, ValueError):
        project.decimals = 4
    project.touch(current_user.username)
    db.session.commit()
    Activity.record(current_user.username, current_user.role,
                    f"Actualizó los datos de «{project.label}»", "sheet")
    return redirect(url_for("validation.menu", project_id=project.id))


@validation_bp.route("/<int:project_id>/modulo/<module_id>")
@login_required
def module(project_id, module_id):
    project = db.get_or_404(Project, project_id)
    mod = MODULE_BY_ID.get(module_id)
    if mod is None:
        abort(404)
    record = ModuleData.query.filter_by(project_id=project.id, module=module_id).first()
    data = record.get_data() if record else {}
    info = {
        "version": record.version if record else 0,
        "updated_by": record.updated_by if record else None,
        "updated_at": record.updated_at.isoformat() + "Z" if record and record.updated_at else None,
    }
    return render_template("validation/module.html",
                           project=project, module=mod, saved=data, info=info)


@validation_bp.route("/<int:project_id>/api/<module_id>/compute", methods=["POST"])
@login_required
def api_compute(project_id, module_id):
    db.get_or_404(Project, project_id)
    if module_id not in MODULE_BY_ID:
        abort(404)
    payload = request.get_json(silent=True) or {}
    project = db.session.get(Project, project_id)
    return jsonify(compute(module_id, payload, project))


@validation_bp.route("/<int:project_id>/api/<module_id>/save", methods=["POST"])
@login_required
def api_save(project_id, module_id):
    project = db.get_or_404(Project, project_id)
    if module_id not in MODULE_BY_ID:
        abort(404)
    body = request.get_json(silent=True) or {}
    payload = body.get("payload", {})
    base_version = int(body.get("base_version") or 0)
    force = bool(body.get("force"))

    record = ModuleData.query.filter_by(project_id=project.id, module=module_id).first()

    if record and not force and base_version and record.version > base_version:
        # Alguien más guardó después de que esta persona abrió el módulo.
        return jsonify({
            "conflict": True,
            "updated_by": record.updated_by,
            "version": record.version,
        }), 409

    if record is None:
        record = ModuleData(project_id=project.id, module=module_id)
        db.session.add(record)
    record.set_data(payload)
    record.updated_by = current_user.username
    project.touch(current_user.username)
    db.session.commit()

    Activity.record(current_user.username, current_user.role,
                    f"Guardó «{MODULE_BY_ID[module_id].name}» en «{project.label}»", "data")
    return jsonify({"saved": True, "version": record.version,
                    "updated_by": record.updated_by})


@validation_bp.route("/<int:project_id>/api/import-excel", methods=["POST"])
@login_required
def import_excel(project_id):
    db.get_or_404(Project, project_id)
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"error": "No se recibió ningún archivo."}), 400
    if os.path.splitext(file.filename)[1].lower() not in ALLOWED_EXCEL:
        return jsonify({"error": "Formato no válido. Sube un archivo .xlsx o .xlsm."}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True, read_only=True)
    except Exception:
        return jsonify({"error": "No se pudo leer el archivo de Excel."}), 400

    tables = []
    for ws in wb.worksheets[:10]:
        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 200:
                break
            data.append(["" if v is None else str(v) for v in row[:50]])
        while data and not any(c.strip() for c in data[-1]):
            data.pop()
        if not data:
            continue
        headers = data[0]
        if any(h.strip() for h in headers):
            body = data[1:]
        else:
            headers = [f"Columna {j + 1}" for j in range(len(headers))]
            body = data
        width = max([len(headers)] + [len(r) for r in body]) if (headers or body) else 0
        headers = (headers + [""] * width)[:width]
        body = [(r + [""] * width)[:width] for r in body]
        grid = [(r + [""] * width)[:width] for r in data]
        tables.append({"title": ws.title, "headers": headers, "rows": body, "grid": grid})
    wb.close()

    if not tables:
        return jsonify({"error": "El archivo no contenía datos legibles."}), 400
    Activity.record(current_user.username, current_user.role,
                    f"Importó datos de Excel ({file.filename})", "data")
    return jsonify({"tables": tables})


@validation_bp.route("/<int:project_id>/informe")
@login_required
def report(project_id):
    project = db.get_or_404(Project, project_id)
    return render_template("validation/report.html", project=project,
                           report=build_report(project))


@validation_bp.route("/<int:project_id>/informe.pdf")
@login_required
def report_pdf(project_id):
    project = db.get_or_404(Project, project_id)
    Activity.record(current_user.username, current_user.role,
                    f"Descargó el informe PDF de «{project.label}»", "sheet")
    pdf = render_pdf(project)
    name = f"informe_{(project.code or 'hoja').replace(' ', '_')}.pdf"
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=name)
